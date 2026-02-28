"""Document processing pipeline — parse, chunk, embed, store.

Adapted from Metis_2's document_pipeline.py and chunking_service.py.
Supports: PDF, DOCX, XLSX, TXT, MD, HTML, CSV.
Chunking: recursive character splitting with configurable size/overlap.
"""

import csv
import io
import logging
import re
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

# ── File type detection ────────────────────────────────────────────────

SUPPORTED_EXTENSIONS: set[str] = {
    ".pdf", ".docx", ".xlsx", ".txt", ".md", ".html", ".htm", ".csv",
}


def is_supported(filename: str) -> bool:
    return Path(filename).suffix.lower() in SUPPORTED_EXTENSIONS


# ── Parsing ────────────────────────────────────────────────────────────

async def parse_document(file_path: str, filename: str) -> list[dict[str, Any]]:
    """Parse a file into a list of ``{text, metadata}`` page-dicts.

    Metadata includes ``page_number`` when available.
    Raises ``ValueError`` for unsupported or empty files.
    """
    ext = Path(filename).suffix.lower()

    if ext == ".pdf":
        return _parse_pdf(file_path)
    elif ext == ".docx":
        return _parse_docx(file_path)
    elif ext == ".xlsx":
        return _parse_xlsx(file_path)
    elif ext in {".txt", ".md"}:
        return _parse_text(file_path)
    elif ext in {".html", ".htm"}:
        return _parse_html(file_path)
    elif ext == ".csv":
        return _parse_csv(file_path)
    else:
        raise ValueError(f"Unsupported file type: {ext}")


def _parse_pdf(file_path: str) -> list[dict[str, Any]]:
    """Extract text from PDF, with OCR fallback for scanned pages."""
    try:
        from pypdf import PdfReader
    except ImportError:
        raise ImportError("pypdf is required for PDF parsing. Install it with: pip install pypdf")

    reader = PdfReader(file_path)
    pages: list[dict[str, Any]] = []

    for idx, page in enumerate(reader.pages):
        text = (page.extract_text() or "").strip()

        # OCR fallback for scanned / image-heavy pages
        if len(text) < 50:
            ocr_text = _ocr_page(file_path, idx)
            if ocr_text:
                text = ocr_text

        if text:
            pages.append({"text": text, "metadata": {"page_number": idx + 1}})

    if not pages:
        raise ValueError("PDF contains no extractable text")
    return pages


def _ocr_page(file_path: str, page_idx: int) -> str | None:
    """Run OCR on a single page via pytesseract. Returns None on failure."""
    try:
        import pdf2image  # type: ignore
        import pytesseract  # type: ignore

        images = pdf2image.convert_from_path(file_path, first_page=page_idx + 1, last_page=page_idx + 1)
        if images:
            return pytesseract.image_to_string(images[0]).strip()
    except Exception:
        logger.debug("OCR unavailable for page %d of %s", page_idx, file_path, exc_info=True)
    return None


def _parse_docx(file_path: str) -> list[dict[str, Any]]:
    try:
        from docx import Document as DocxDocument  # type: ignore
    except ImportError:
        raise ImportError("python-docx is required. Install with: pip install python-docx")

    doc = DocxDocument(file_path)
    full_text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    if not full_text.strip():
        raise ValueError("DOCX contains no text content")
    return [{"text": full_text, "metadata": {"page_number": None}}]


def _parse_xlsx(file_path: str) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError:
        raise ImportError("openpyxl is required. Install with: pip install openpyxl")

    wb = load_workbook(file_path, read_only=True, data_only=True)
    pages: list[dict[str, Any]] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows: list[str] = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(cells):
                rows.append(" | ".join(cells))
        if rows:
            text = f"Sheet: {sheet_name}\n" + "\n".join(rows)
            pages.append({"text": text, "metadata": {"page_number": None, "sheet_name": sheet_name}})
    wb.close()
    if not pages:
        raise ValueError("XLSX contains no data")
    return pages


def _parse_text(file_path: str) -> list[dict[str, Any]]:
    text = Path(file_path).read_text(encoding="utf-8", errors="replace").strip()
    if not text:
        raise ValueError("Text file is empty")
    return [{"text": text, "metadata": {"page_number": None}}]


def _parse_html(file_path: str) -> list[dict[str, Any]]:
    raw = Path(file_path).read_text(encoding="utf-8", errors="replace")
    # Strip tags for a simple text extraction
    clean = re.sub(r"<[^>]+>", " ", raw)
    clean = re.sub(r"\s+", " ", clean).strip()
    if not clean:
        raise ValueError("HTML file contains no text")
    return [{"text": clean, "metadata": {"page_number": None}}]


def _parse_csv(file_path: str) -> list[dict[str, Any]]:
    text_lines: list[str] = []
    with open(file_path, encoding="utf-8", errors="replace") as f:
        reader = csv.reader(f)
        for row in reader:
            if any(cell.strip() for cell in row):
                text_lines.append(" | ".join(row))
    if not text_lines:
        raise ValueError("CSV file is empty")
    return [{"text": "\n".join(text_lines), "metadata": {"page_number": None}}]


# ── Chunking ───────────────────────────────────────────────────────────

def chunk_text(
    text: str,
    chunk_size: int = 512,
    chunk_overlap: int = 50,
    separators: list[str] | None = None,
) -> list[str]:
    """Recursive character splitting with configurable size and overlap.

    Adapted from Metis_2's ChunkingService with adaptive sizing.
    """
    if not text:
        return []

    if separators is None:
        separators = ["\n\n", "\n", ". ", " ", ""]

    # Adaptive chunk sizing based on text length (from Metis_2)
    total_length = len(text)
    if total_length < 1500:
        chunk_size = min(chunk_size, 800)
        chunk_overlap = min(chunk_overlap, 100)
    elif total_length < 3000:
        chunk_size = min(chunk_size, 1200)
        chunk_overlap = min(chunk_overlap, 200)

    return _recursive_split(text, chunk_size, chunk_overlap, separators)


def _recursive_split(
    text: str,
    chunk_size: int,
    chunk_overlap: int,
    separators: list[str],
) -> list[str]:
    """Split text recursively, trying each separator in order."""
    if len(text) <= chunk_size:
        stripped = text.strip()
        return [stripped] if stripped else []

    # Find the best separator for this text
    separator = separators[-1]
    for sep in separators:
        if sep in text:
            separator = sep
            break

    splits = text.split(separator) if separator else list(text)
    chunks: list[str] = []
    current_chunk: list[str] = []
    current_length = 0

    for piece in splits:
        piece_len = len(piece) + (len(separator) if current_chunk else 0)

        if current_length + piece_len > chunk_size and current_chunk:
            chunk_text_str = separator.join(current_chunk).strip()
            if chunk_text_str:
                chunks.append(chunk_text_str)

            # Keep overlap
            overlap_pieces: list[str] = []
            overlap_len = 0
            for p in reversed(current_chunk):
                if overlap_len + len(p) > chunk_overlap:
                    break
                overlap_pieces.insert(0, p)
                overlap_len += len(p)
            current_chunk = overlap_pieces
            current_length = overlap_len

        current_chunk.append(piece)
        current_length += piece_len

    # Final chunk
    if current_chunk:
        chunk_text_str = separator.join(current_chunk).strip()
        if chunk_text_str:
            chunks.append(chunk_text_str)

    # Quality filter: discard very short chunks
    MIN_CHUNK_LENGTH = 30
    quality_chunks = [c for c in chunks if len(c) >= MIN_CHUNK_LENGTH]

    return quality_chunks


def chunk_pages(
    pages: list[dict[str, Any]],
    chunk_size: int = 512,
    chunk_overlap: int = 50,
) -> list[dict[str, Any]]:
    """Chunk parsed pages and preserve page-number metadata.

    Returns list of ``{text, metadata}`` where metadata includes
    ``page_number`` and ``chunk_index``.
    """
    all_chunks: list[dict[str, Any]] = []
    global_idx = 0

    for page in pages:
        raw_text = page["text"]
        page_meta = page.get("metadata", {})
        page_number = page_meta.get("page_number")

        text_chunks = chunk_text(raw_text, chunk_size=chunk_size, chunk_overlap=chunk_overlap)
        for chunk in text_chunks:
            all_chunks.append(
                {
                    "text": chunk,
                    "metadata": {
                        **page_meta,
                        "page_number": page_number,
                        "chunk_index": global_idx,
                    },
                }
            )
            global_idx += 1

    return all_chunks
